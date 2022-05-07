from jproperties import Properties
import openpyxl
'''
Read from txt files.
'''

file=open('file.txt', 'r')
for each in file:
    print(each)
print("**********")
with open("file.txt") as file:
    data=file.read()
    print(data)

'''
Read from property files
'''
configs=Properties()
with open("example.properties", 'rb') as read_prop:
    configs.load(read_prop)
prop_view = configs.items()
print(type(prop_view))
print("*********Way 1 of printing properties***********")
for item in prop_view:
    print(item)
print("*********Way 2 of printing properties***********")
for item in prop_view:
    print(item[0], '=', item[1].data)
print("*********Way 3 of printing properties***********")
print(configs.get("DB_User"))
print(f'Database User: {configs.get("DB_User").data}')
print(f'Database Password: {configs["DB_PWD"].data}')
print(f'Properties Count: {len(configs)}')

'''
# Reading an excel file using Python
'''
book = openpyxl.load_workbook('ExcelExample.xlsx')
sheet = book.active
a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)
print(a1.value)
print(a2.value)
print(a3.value)

'''
Reading an excel file Method 2
'''
print("**********Method 2 of reading excel file**********")
book = openpyxl.load_workbook('ExcelExample.xlsx')
sheet = book.active
cells = sheet['A1': 'B6']
for c1, c2 in cells:
    print("{0:2} {1:2}".format(c1.value, c2.value))

    